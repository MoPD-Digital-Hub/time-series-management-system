from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.core.paginator import Paginator
from django.db.models import Q
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.utils import timezone
from ..models import CustomUser, CategoryAssignment, IndicatorSubmission, DataSubmission
from Base.models import Category, Indicator, DataPoint, Month, MonthData, Quarter, QuarterData, AnnualData, KPIRecord
from ..serializers import (
    CustomUserSerializer, CategoryAssignmentSerializer,
    IndicatorSubmissionSerializer, DataSubmissionSerializer,
    UserManagementStatsSerializer, UnassignedCategorySerializer
)
from ethiopian_date_converter.ethiopian_date_convertor import to_ethiopian, to_gregorian, EthDate
import secrets
from ..models import CustomUser as UM_CustomUser
import csv
import os
from django.http import HttpResponse
import tablib
try:
    import openpyxl
except Exception:
    openpyxl = None

from io import TextIOWrapper

from rest_framework.views import APIView


@api_view(['GET'])
def user_management_stats_api(request):
    """Get user management dashboard statistics"""
    try:
        user = request.user
        if user.is_staff or user.is_superuser:
            stats = {
                'total_users': CustomUser.objects.count(),
                'active_users': CustomUser.objects.filter(is_active=True).count(),
                'category_managers': CustomUser.objects.filter(is_category_manager=True).count(),
                'importers': CustomUser.objects.filter(is_importer=True).count(),
                'pending_indicator_submissions': IndicatorSubmission.objects.filter(status='pending').count(),
                'pending_data_submissions': DataSubmission.objects.filter(status='pending').count(),
            }
        else:
            # For category managers
            managed_importers = CustomUser.objects.filter(manager=user)
            assigned_categories = CategoryAssignment.objects.filter(manager=user).values_list('category_id', flat=True)
            
            stats = {
                'total_users': managed_importers.count() + 1, # Importers + self
                'active_users': managed_importers.filter(is_active=True).count() + 1,
                'category_managers': 1, # Just self
                'importers': managed_importers.count(),
                'pending_indicator_submissions': IndicatorSubmission.objects.filter(
                    Q(status='pending') & 
                    (Q(submitted_by__in=managed_importers) | Q(indicator__for_category__in=assigned_categories))
                ).distinct().count(),
                'pending_data_submissions': DataSubmission.objects.filter(
                    Q(status='pending') & 
                    (Q(submitted_by__in=managed_importers) | Q(indicator__for_category__in=assigned_categories))
                ).distinct().count(),
            }
        
        serializer = UserManagementStatsSerializer(stats)
        return Response(serializer.data)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def users_list_api(request):
    """Get paginated list of users with filtering"""
    search_query = request.GET.get('search', '')
    role_filter = request.GET.get('role', '')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    
    users = CustomUser.objects.all()
    
    if search_query:
        users = users.filter(
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(username__icontains=search_query)
        )
    
    if role_filter:
        if role_filter == 'category_manager':
            users = users.filter(is_category_manager=True)
        elif role_filter == 'importer':
            users = users.filter(is_importer=True)
        elif role_filter == 'admin':
            users = users.filter(is_staff=True)
    
    # Pagination
    paginator = Paginator(users, page_size)
    users_page = paginator.get_page(page)
    
    serializer = CustomUserSerializer(users_page.object_list, many=True)
    
    return Response({
        'results': serializer.data,
        'pagination': {
            'current_page': users_page.number,
            'total_pages': paginator.num_pages,
            'total_count': paginator.count,
            'has_next': users_page.has_next(),
            'has_previous': users_page.has_previous(),
            'start_index': users_page.start_index(),
            'end_index': users_page.end_index(),
        }
    })


@api_view(['GET'])
def indicator_submissions_api(request):
    """Get paginated list of indicator submissions with filtering"""
    status_filter = request.GET.get('status', '')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    
    submissions = IndicatorSubmission.objects.select_related(
        'indicator', 'submitted_by', 'verified_by'
    ).order_by('-submitted_at')
    
    if not (request.user.is_staff or request.user.is_superuser):
        managed_importers = CustomUser.objects.filter(manager=request.user)
        assigned_categories = CategoryAssignment.objects.filter(manager=request.user).values_list('category_id', flat=True)
        submissions = submissions.filter(
            Q(submitted_by__in=managed_importers) |
            Q(indicator__for_category__in=assigned_categories)
        ).distinct()

    if status_filter:
        submissions = submissions.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(submissions, page_size)
    submissions_page = paginator.get_page(page)
    
    serializer = IndicatorSubmissionSerializer(submissions_page.object_list, many=True)
    
    return Response({
        'results': serializer.data,
        'pagination': {
            'current_page': submissions_page.number,
            'total_pages': paginator.num_pages,
            'total_count': paginator.count,
            'has_next': submissions_page.has_next(),
            'has_previous': submissions_page.has_previous(),
            'start_index': submissions_page.start_index(),
            'end_index': submissions_page.end_index(),
        }
    })


@api_view(['GET'])
def data_submissions_api(request):
    """Get paginated list of data submissions with filtering"""
    status_filter = request.GET.get('status', '')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    
    submissions = DataSubmission.objects.select_related(
        'indicator', 'submitted_by', 'verified_by'
    ).order_by('-submitted_at')
    
    if not (request.user.is_staff or request.user.is_superuser):
        managed_importers = CustomUser.objects.filter(manager=request.user)
        assigned_categories = CategoryAssignment.objects.filter(manager=request.user).values_list('category_id', flat=True)
        submissions = submissions.filter(
            Q(submitted_by__in=managed_importers) |
            Q(indicator__for_category__in=assigned_categories)
        ).distinct()

    if status_filter:
        submissions = submissions.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(submissions, page_size)
    submissions_page = paginator.get_page(page)
    
    serializer = DataSubmissionSerializer(submissions_page.object_list, many=True)
    
    return Response({
        'results': serializer.data,
        'pagination': {
            'current_page': submissions_page.number,
            'total_pages': paginator.num_pages,
            'total_count': paginator.count,
            'has_next': submissions_page.has_next(),
            'has_previous': submissions_page.has_previous(),
            'start_index': submissions_page.start_index(),
            'end_index': submissions_page.end_index(),
        }
    })


@api_view(['GET'])
def category_assignments_api(request):
    """Get paginated list of category assignments"""
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    
    assignments = CategoryAssignment.objects.select_related(
        'manager', 'category'
    ).order_by('category__name_ENG')
    
    # Pagination
    paginator = Paginator(assignments, page_size)
    assignments_page = paginator.get_page(page)
    
    serializer = CategoryAssignmentSerializer(assignments_page.object_list, many=True)
    
    return Response({
        'results': serializer.data,
        'pagination': {
            'current_page': assignments_page.number,
            'total_pages': paginator.num_pages,
            'total_count': paginator.count,
            'has_next': assignments_page.has_next(),
            'has_previous': assignments_page.has_previous(),
            'start_index': assignments_page.start_index(),
            'end_index': assignments_page.end_index(),
        }
    })


@api_view(['GET'])
def unassigned_categories_api(request):
    """Get list of categories not assigned to any manager with optional filtering"""
    topic_id = request.GET.get('topic_id')
    search = request.GET.get('search')
    
    assigned_category_ids = CategoryAssignment.objects.values_list('category_id', flat=True)
    unassigned_categories = Category.objects.filter(topic__is_initiative=False).exclude(id__in=assigned_category_ids).prefetch_related('indicators')
    
    if topic_id:
        unassigned_categories = unassigned_categories.filter(topic_id=topic_id)
        
    if search:
        unassigned_categories = unassigned_categories.filter(
            Q(name_ENG__icontains=search) | Q(name_AMH__icontains=search)
        )
        
    serializer = UnassignedCategorySerializer(unassigned_categories, many=True)
    return Response(serializer.data)


@api_view(['POST'])
def create_importer_api(request):
    """Allow a category manager to create an importer, or admin to create a category manager."""
    if not request.user.is_authenticated:
        return Response({'error': 'Authentication required.'}, status=status.HTTP_403_FORBIDDEN)
    
    # Allow admins to create category managers, or category managers to create importers
    is_admin = request.user.is_staff
    is_manager = request.user.is_category_manager
    
    if not (is_admin or is_manager):
        return Response({'error': 'Access denied. Only admins or category managers can create users.'}, status=status.HTTP_403_FORBIDDEN)

    # support both JSON and multipart/form-data (photo upload)
    email = (request.data.get('email') or '').strip().lower()
    first_name = (request.data.get('first_name') or '').strip()
    last_name = (request.data.get('last_name') or '').strip()
    password = request.data.get('password') or None
    photo = request.FILES.get('photo')
    # support multiple category IDs (e.g. from select multiple)
    assigned_categories = request.data.getlist('assigned_categories[]') or request.data.getlist('assigned_categories') or []

    if not email or '@' not in email:
        return Response({'error': 'A valid email is required.'}, status=status.HTTP_400_BAD_REQUEST)

    if UM_CustomUser.objects.filter(email=email).exists():
        return Response({'error': 'A user with that email already exists.'}, status=status.HTTP_400_BAD_REQUEST)

    # generate password if not provided
    if not password:
        password = secrets.token_urlsafe(8)

    # derive a username from email (ensure unique)
    base_username = email.split('@')[0][:30]
    username = base_username
    suffix = 1
    while UM_CustomUser.objects.filter(username=username).exists():
        username = f"{base_username}{suffix}"
        suffix += 1

    try:
        user = UM_CustomUser.objects.create_user(email=email, username=username, first_name=first_name, last_name=last_name, password=password)
        
        # If admin is creating, make it a category manager; if manager is creating, make it an importer
        if is_admin:
            user.is_category_manager = True
        else:
            user.is_importer = True
            user.manager = request.user
        
        # Ensure user is active (from is_active field in form if provided, otherwise default to True)
        is_active = request.data.get('is_active')
        if is_active is not None:
            user.is_active = str(is_active).lower() in ('true', '1', 'yes')
        else:
            user.is_active = True  # Default to active for new users
        
        if photo:
            user.photo = photo
        user.save()

        # Create Category Assignments
        created_assignments = []
        for cat_id in assigned_categories:
            try:
                cat_id = int(cat_id)
                # Admins can assign any category; managers can only assign their own categories
                if is_admin or CategoryAssignment.objects.filter(manager=request.user, category_id=cat_id).exists():
                    CategoryAssignment.objects.get_or_create(manager=user, category_id=cat_id)
                    created_assignments.append(cat_id)
            except (ValueError, TypeError, Category.DoesNotExist):
                continue

        user_type = 'Category Manager' if is_admin else 'Importer'
        resp = {'message': f'{user_type} created', 'id': user.id, 'email': user.email}
        if password:
            resp['password'] = password
        resp['assigned_categories'] = created_assignments
        return Response(resp, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST', 'PATCH'])
@permission_classes([IsAuthenticated])
def update_user_api(request, pk):
    """Update user information and category assignments."""
    user = get_object_or_404(CustomUser, pk=pk)
    
    # Permission check: superuser/admin can edit anyone, manager can edit their importers
    if not request.user.is_superuser and not request.user.is_staff:
        if not (request.user.is_category_manager and user.manager == request.user):
            return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

    email = (request.data.get('email') or '').strip().lower()
    first_name = (request.data.get('first_name') or '').strip()
    last_name = (request.data.get('last_name') or '').strip()
    password = request.data.get('password')
    is_active = request.data.get('is_active')
    photo = request.FILES.get('photo')
    assigned_categories = request.data.getlist('assigned_categories[]') or request.data.getlist('assigned_categories') or []

    if email and email != user.email:
        if CustomUser.objects.filter(email=email).exclude(pk=pk).exists():
            return Response({'error': 'A user with that email already exists.'}, status=status.HTTP_400_BAD_REQUEST)
        user.email = email
        user.username = email.split('@')[0] # sync username roughly or keep it same
    
    if first_name: user.first_name = first_name
    if last_name: user.last_name = last_name
    if password: user.set_password(password)
    if is_active is not None:
        user.is_active = str(is_active).lower() in ('true', '1', 'yes')
    if photo: user.photo = photo
    
    user.save()

    # Sync Category Assignments
    # For simplicity, we clear and recreate or just sync.
    # We only allow syncing categories that the CURRENT manager (request.user) has access to.
    manager_accessible_cats = []
    if request.user.is_superuser or request.user.is_staff:
        manager_accessible_cats = list(Category.objects.filter(topic__is_initiative=False).values_list('id', flat=True))
    else:
        manager_accessible_cats = list(CategoryAssignment.objects.filter(manager=request.user, category__topic__is_initiative=False).values_list('category_id', flat=True))

    valid_new_cat_ids = []
    for cid in assigned_categories:
        try:
            cid = int(cid)
            if cid in manager_accessible_cats:
                valid_new_cat_ids.append(cid)
        except: continue

    # Remove assignments that are no longer in the list (but only those the manager had power over)
    CategoryAssignment.objects.filter(manager=user, category_id__in=manager_accessible_cats).exclude(category_id__in=valid_new_cat_ids).delete()
    
    # Add new assignments
    for cid in valid_new_cat_ids:
        CategoryAssignment.objects.get_or_create(manager=user, category_id=cid)

    return Response({
        'message': 'User updated successfully',
        'id': user.id,
        'email': user.email,
        'assigned_categories': valid_new_cat_ids
    })


@api_view(['GET'])
def recent_submissions_api(request):
    """Get recent indicator and data submissions for dashboard"""
    limit = int(request.GET.get('limit', 5))
    
    recent_indicator_submissions = IndicatorSubmission.objects.select_related(
        'indicator', 'submitted_by', 'verified_by'
    ).order_by('-submitted_at')
    
    recent_data_submissions = DataSubmission.objects.select_related(
        'indicator', 'submitted_by', 'verified_by'
    ).order_by('-submitted_at')

    if not (request.user.is_staff or request.user.is_superuser):
        managed_importers = CustomUser.objects.filter(manager=request.user)
        assigned_categories = CategoryAssignment.objects.filter(manager=request.user).values_list('category_id', flat=True)
        
        recent_indicator_submissions = recent_indicator_submissions.filter(
            Q(submitted_by__in=managed_importers) |
            Q(indicator__for_category__in=assigned_categories)
        ).distinct()
        
        recent_data_submissions = recent_data_submissions.filter(
            Q(submitted_by__in=managed_importers) |
            Q(indicator__for_category__in=assigned_categories)
        ).distinct()

    recent_indicator_submissions = recent_indicator_submissions[:limit]
    recent_data_submissions = recent_data_submissions[:limit]
    
    indicator_serializer = IndicatorSubmissionSerializer(recent_indicator_submissions, many=True)
    data_serializer = DataSubmissionSerializer(recent_data_submissions, many=True)
    
    return Response({
        'indicator_submissions': indicator_serializer.data,
        'data_submissions': data_serializer.data,
    })


@api_view(['GET'])
def recent_table_data_submissions_api(request):
    """
    Get recent cell-level edits (unverified data points) for the dashboard.
    """
    limit = int(request.GET.get('limit', 10))
    
    # We fetch unverified data from all data models
    # and combine them into a single list, sorted by created_at/date.
    
    q_filter = Q(is_verified=False)
    if not (request.user.is_staff or request.user.is_superuser):
        if request.user.is_importer:
            # Importers see their own pending data
            q_filter &= Q(submitted_by=request.user)
            # Managers see managed importers or assigned categories
            managed_importers = CustomUser.objects.filter(manager=request.user)
            assigned_categories = CategoryAssignment.objects.filter(manager=request.user, category__topic__is_initiative=False).values_list('category_id', flat=True)
            q_filter &= (Q(submitted_by__in=managed_importers) | Q(indicator__for_category__in=assigned_categories))
        else:
            # Other roles (if any) see nothing by default
            return Response([])
    
    # Global filter for initiatives
    q_filter &= Q(indicator__for_category__topic__is_initiative=False)

    # Fetch from models
    annual = AnnualData.objects.filter(q_filter).select_related('indicator', 'submitted_by', 'for_datapoint').order_by('-created_at')[:limit]
    monthly = MonthData.objects.filter(q_filter).select_related('indicator', 'submitted_by', 'for_month', 'for_datapoint').order_by('-created_at')[:limit]
    quarterly = QuarterData.objects.filter(q_filter).select_related('indicator', 'submitted_by', 'for_quarter', 'for_datapoint').order_by('-created_at')[:limit]
    kpi = KPIRecord.objects.filter(q_filter).select_related('indicator', 'submitted_by').order_by('-created_at')[:limit]

    results = []
    
    for item in annual:
        results.append({
            'id': item.id,
            'type': 'Annual',
            'indicator_name': item.indicator.title_ENG if item.indicator else "N/A",
            'submitted_by': item.submitted_by.email if item.submitted_by else "N/A",
            'value': item.performance,
            'period': item.for_datapoint.year_EC if item.for_datapoint else "N/A",
            'created_at': item.created_at
        })
    
    for item in monthly:
        results.append({
            'id': item.id,
            'type': 'Monthly',
            'indicator_name': item.indicator.title_ENG if item.indicator else "N/A",
            'submitted_by': item.submitted_by.email if item.submitted_by else "N/A",
            'value': item.performance,
            'period': f"{item.for_month.month_ENG} {item.for_datapoint.year_EC}" if item.for_month and item.for_datapoint else "N/A",
            'created_at': item.created_at
        })
        
    for item in quarterly:
        results.append({
            'id': item.id,
            'type': 'Quarterly',
            'indicator_name': item.indicator.title_ENG if item.indicator else "N/A",
            'submitted_by': item.submitted_by.email if item.submitted_by else "N/A",
            'value': item.performance,
            'period': f"{item.for_quarter.title_ENG} {item.for_datapoint.year_EC}" if item.for_quarter and item.for_datapoint else "N/A",
            'created_at': item.created_at
        })
        
    for item in kpi:
        results.append({
            'id': item.id,
            'type': item.record_type.title(),
            'indicator_name': item.indicator.title_ENG if item.indicator else "N/A",
            'submitted_by': item.submitted_by.email if item.submitted_by else "N/A",
            'value': str(item.performance) if item.performance is not None else "N/A",
            'period': item.date.strftime('%Y-%m-%d') if item.date else "N/A",
            'created_at': item.created_at
        })

    # Sort results by created_at (if available) or period/id
    # Ensure created_at is handled safely
    results.sort(key=lambda x: (x['created_at'].timestamp() if x['created_at'] else 0), reverse=True)
    
    return Response(results[:limit])


@api_view(['POST'])
def approve_all_submissions_api(request):
    """Approve all pending submissions of a specific type (indicator or data)"""
    submission_type = request.data.get('type')
    
    if submission_type not in ['indicator', 'data']:
        return Response({'error': 'Invalid submission type'}, status=status.HTTP_400_BAD_REQUEST)
    
    submissions = None
    if submission_type == 'indicator':
        submissions = IndicatorSubmission.objects.filter(status='pending')
    else:
        submissions = DataSubmission.objects.filter(status='pending')
        
    if not (request.user.is_staff or request.user.is_superuser):
        managed_importers = CustomUser.objects.filter(manager=request.user)
        assigned_categories = CategoryAssignment.objects.filter(manager=request.user).values_list('category_id', flat=True)
        
        if submission_type == 'indicator':
            submissions = submissions.filter(
                Q(submitted_by__in=managed_importers) | 
                Q(indicator__for_category__in=assigned_categories)
            ).distinct()
        else:
            submissions = submissions.filter(
                Q(submitted_by__in=managed_importers) | 
                Q(indicator__for_category__in=assigned_categories)
            ).distinct()

    count = submissions.count()
    for submission in submissions:
        submission.status = 'approved'
        submission.verified_by = request.user
        submission.verified_at = timezone.now()
        submission.save()
        
        if submission_type == 'indicator' and submission.indicator:
            try:
                ind = submission.indicator
                ind.is_verified = True
                ind.save(update_fields=['is_verified'])
            except: pass
        elif submission_type == 'data':
            try:
                _import_data_submission_to_db(submission)
            except: pass
            
    return Response({'message': f'Successfully approved {count} {submission_type} submissions.'})


@api_view(['POST'])
def approve_submission_api(request):
    """Approve a submission (indicator or data)"""
    submission_type = request.data.get('type')
    submission_id = request.data.get('id')
    
    if submission_type == 'indicator':
        submission = get_object_or_404(IndicatorSubmission, id=submission_id)
    elif submission_type == 'data':
        submission = get_object_or_404(DataSubmission, id=submission_id)
    else:
        return Response({'error': 'Invalid submission type'}, status=status.HTTP_400_BAD_REQUEST)
    
    submission.status = 'approved'
    submission.verified_by = request.user
    submission.verified_at = timezone.now()
    submission.save()
    # If this is an indicator submission, mark the indicator as verified
    if submission_type == 'indicator' and hasattr(submission, 'indicator') and submission.indicator:
        try:
            ind = submission.indicator
            ind.is_verified = True
            ind.save(update_fields=['is_verified'])
        except Exception:
            # don't block approval if marking fails; approval already saved
            pass
    
    if submission_type == 'indicator':
        serializer = IndicatorSubmissionSerializer(submission)
    else:
        # attempt to import data into DB when a data submission is approved
        import_result = None
        try:
            import_result = _import_data_submission_to_db(submission)
        except Exception as e:
            # don't fail the approval if import fails; include error in response
            import_result = {'error': str(e)}
        serializer = DataSubmissionSerializer(submission)
    
    out = serializer.data
    if submission_type == 'data':
        # keep serialized submission fields at top-level for backwards compatibility
        # and attach import_result as an extra key
        try:
            out = dict(serializer.data)
        except Exception:
            out = serializer.data
        out['import_result'] = import_result
    return Response(out)


def _import_data_submission_to_db(submission: DataSubmission):


    result = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}

    # ensure file path is available
    ffield = submission.data_file
    if not ffield:
        raise ValueError('No uploaded file attached to submission')

    file_path = ffield.path if hasattr(ffield, 'path') else None
    if not file_path or not os.path.exists(file_path):
        raise ValueError('Uploaded file not found on disk')

    # helper to parse rows from CSV or Excel
    def iter_rows_from_file(path):
        lower = path.lower()
        if lower.endswith('.csv'):
            with open(path, newline='', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                for r in reader:
                    yield r
        elif lower.endswith(('.xls', '.xlsx')):
            if not openpyxl:
                raise RuntimeError('openpyxl required to parse Excel files but is not installed')
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.rows)
            if not rows:
                return
            headers = [str(c.value).strip() if c.value is not None else '' for c in rows[0]]
            for row in rows[1:]:
                obj = {}
                for h, cell in zip(headers, row):
                    obj[h] = cell.value
                yield obj
        else:
            raise RuntimeError('Unsupported file type')

    # indicator identification
    submission_indicator = submission.indicator
    indicator_code = (submission_indicator.code or '').strip().lower() if submission_indicator else None

    # parse and import
    for i, raw in enumerate(iter_rows_from_file(file_path), start=1):
        try:
            # normalize keys to lowercase
            row = { (k or '').strip().lower(): (v if v is not None else '') for k,v in raw.items() }

            # Determine which indicator this row belongs to
            row_indicator_code = (row.get('indicator') or '').strip().lower()
            
            target_indicator = None
            if submission_indicator:
                # Single mode: only import rows matching the submission's indicator
                if row_indicator_code and indicator_code and row_indicator_code != indicator_code:
                    continue
                target_indicator = submission_indicator
            else:
                # Bulk mode (Multiple mode): look up indicator from row
                if not row_indicator_code:
                    # Skip rows with no indicator in bulk mode
                    continue
                try:
                    target_indicator = Indicator.objects.get(code__iexact=row_indicator_code)
                except Indicator.DoesNotExist:
                    result['skipped'] += 1
                    result['errors'].append({'row': i, 'error': f'Unknown indicator code: {row_indicator_code}'})
                    continue

            # Check if it's "Wide" format
            is_annual_wide = False
            year_cols = []
            for k in row.keys():
                if k.isdigit() and len(k) == 4:
                    year_cols.append(k)
            if year_cols:
                is_annual_wide = True

            is_period_wide = ('for_datapoint' in row) and ('for_quarter' in row or 'for_month' in row)
            
            if is_annual_wide:
                # Wide format (Annual only - years as columns)
                for year_str in year_cols:
                    perf_raw = row.get(year_str)
                    if perf_raw in (None, ''):
                        continue
                    try:
                        performance = float(perf_raw)
                        datapoint, _ = DataPoint.objects.get_or_create(year_EC=year_str)
                        obj, created = AnnualData.objects.update_or_create(
                            indicator=target_indicator,
                            for_datapoint=datapoint,
                            defaults={'performance': performance, 'is_verified': True}
                        )
                        if created: result['created'] += 1
                        else: result['updated'] += 1
                    except Exception as e:
                        result['errors'].append({'row': i, 'error': f'Invalid value for {year_str}: {perf_raw}'})
                continue

            if is_period_wide:
                # Wide-indicator format: for_datapoint, for_quarter/for_month, IND1, IND2...
                year_raw = (row.get('for_datapoint') or '').strip()
                if not year_raw:
                    continue
                
                datapoint, _ = DataPoint.objects.get_or_create(year_EC=str(year_raw))
                
                # Identify period
                quarter_num = None
                month_num = None
                if 'for_quarter' in row and row.get('for_quarter') != '':
                    try: quarter_num = int(row.get('for_quarter'))
                    except: pass
                if 'for_month' in row and row.get('for_month') != '':
                    try: month_num = int(row.get('for_month'))
                    except: pass
                
                # Identify indicator columns (all except control columns)
                control_cols = ['for_datapoint', 'for_quarter', 'for_month', 'indicator', 'title_eng', 'title_amh']
                for col_name, val in row.items():
                    if col_name in control_cols or val in (None, ''):
                        continue
                    
                    # Match indicator by code or title
                    ind_obj = Indicator.objects.filter(code__iexact=col_name).first()
                    if not ind_obj:
                        ind_obj = Indicator.objects.filter(title_ENG__iexact=col_name).first()
                    
                    if not ind_obj:
                        continue
                    
                    try:
                        performance = float(val)
                        if quarter_num:
                            q_obj, _ = Quarter.objects.get_or_create(number=quarter_num, defaults={'title_ENG': f'Q{quarter_num}', 'title_AMH': f'Q{quarter_num}'})
                            obj, created = QuarterData.objects.update_or_create(
                                indicator=ind_obj,
                                for_datapoint=datapoint,
                                for_quarter=q_obj,
                                defaults={'performance': performance, 'is_verified': True}
                            )
                        elif month_num:
                            m_obj = Month.objects.filter(number=month_num).first()
                            if not m_obj:
                                m_obj = Month.objects.create(number=month_num, month_ENG=str(month_num), month_AMH=str(month_num))
                            obj, created = MonthData.objects.update_or_create(
                                indicator=ind_obj,
                                for_datapoint=datapoint,
                                for_month=m_obj,
                                defaults={'performance': performance, 'is_verified': True}
                            )
                        else:
                            # Assume annual if neither quarter nor month
                            obj, created = AnnualData.objects.update_or_create(
                                indicator=ind_obj,
                                for_datapoint=datapoint,
                                defaults={'performance': performance, 'is_verified': True}
                            )
                        
                        if created: result['created'] += 1
                        else: result['updated'] += 1
                    except Exception as e:
                        result['errors'].append({'row': i, 'error': f'Invalid value for indicator {col_name}: {val}'})
                continue

            # Long format (continues below)
            year = row.get('year_ec') or row.get('year_gc')
            if not year:
                result['skipped'] += 1
                result['errors'].append({'row': i, 'error': 'Missing year_EC/year_GC or year-columns'})
                continue

            # performance
            perf_raw = row.get('performance') or row.get('value') or row.get('amount')
            if perf_raw in (None, ''):
                result['skipped'] += 1
                result['errors'].append({'row': i, 'error': 'Missing performance/value'})
                continue
            try:
                performance = float(perf_raw)
            except Exception:
                result['skipped'] += 1
                result['errors'].append({'row': i, 'error': f'Invalid performance value: {perf_raw}'})
                continue

            # find or create datapoint by year_EC (prefer)
            year_ec = row.get('year_ec') or None
            if not year_ec and row.get('year_gc'):
                try:
                    year_ec = str(int(row.get('year_gc')) - 7)
                except Exception:
                    year_ec = None

            if year_ec:
                datapoint, _ = DataPoint.objects.get_or_create(year_EC=str(year_ec))
            else:
                datapoint, _ = DataPoint.objects.get_or_create(year_EC=str(year))

            # frequency
            if 'month' in row and row.get('month') != '':
                mraw = str(row.get('month')).strip()
                month_obj = None
                try:
                    mnum = int(mraw)
                    month_obj = Month.objects.filter(number=mnum).first()
                except Exception:
                    month_obj = Month.objects.filter(month_ENG__iexact=mraw).first()
                if not month_obj:
                    try:
                        if mraw.isdigit():
                            month_obj = Month.objects.create(number=int(mraw), month_ENG=str(mraw), month_AMH=str(mraw))
                        else:
                            raise ValueError('Unknown month')
                    except Exception:
                        result['skipped'] += 1
                        result['errors'].append({'row': i, 'error': f'Invalid month: {mraw}'})
                        continue

                obj, created = MonthData.objects.update_or_create(
                    indicator=target_indicator,
                    for_datapoint=datapoint,
                    for_month=month_obj,
                    defaults={'performance': performance, 'is_verified': True}
                )
                if created: result['created'] += 1
                else: result['updated'] += 1
                continue

            # daily
            if 'day' in row and row.get('day') != '':
                mraw = str(row.get('month')).strip()
                draw = str(row.get('day')).strip()
                try:
                    month = int(mraw)
                    day = int(draw)
                    # Convert to Gregorian
                    eth_date = EthDate(day, month, int(year_ec))
                    greg_date = to_gregorian(eth_date)
                    
                    obj, created = KPIRecord.objects.update_or_create(
                        indicator=target_indicator,
                        date=greg_date,
                        record_type='daily',
                        defaults={'performance': performance, 'is_verified': True}
                    )
                    if created: result['created'] += 1
                    else: result['updated'] += 1
                    continue
                except Exception as e:
                    result['skipped'] += 1
                    result['errors'].append({'row': i, 'error': f'Invalid daily date: {e}'})
                    continue

            # weekly
            if 'week' in row and row.get('week') != '':
                mraw = str(row.get('month')).strip()
                wraw = str(row.get('week')).strip()
                try:
                    month = int(mraw)
                    week = int(wraw)
                    # Following resource.py logic: day = (week-1)*7 + 1
                    day = ((week - 1) * 7) + 1
                    eth_date = EthDate(day, month, int(year_ec))
                    greg_date = to_gregorian(eth_date)
                    
                    obj, created = KPIRecord.objects.update_or_create(
                        indicator=target_indicator,
                        date=greg_date,
                        record_type='weekly',
                        defaults={'performance': performance, 'is_verified': True}
                    )
                    if created: result['created'] += 1
                    else: result['updated'] += 1
                    continue
                except Exception as e:
                    result['skipped'] += 1
                    result['errors'].append({'row': i, 'error': f'Invalid weekly date: {e}'})
                    continue

            if 'quarter' in row and row.get('quarter') != '':
                qraw = row.get('quarter')
                try:
                    qnum = int(qraw)
                except Exception:
                    result['skipped'] += 1
                    result['errors'].append({'row': i, 'error': f'Invalid quarter: {qraw}'})
                    continue
                quarter_obj, _ = Quarter.objects.get_or_create(number=qnum, defaults={'title_ENG': f'Q{qnum}', 'title_AMH': f'Q{qnum}'})
                obj, created = QuarterData.objects.update_or_create(
                    indicator=target_indicator,
                    for_datapoint=datapoint,
                    for_quarter=quarter_obj,
                    defaults={'performance': performance, 'is_verified': True}
                )
                if created: result['created'] += 1
                else: result['updated'] += 1
                continue

            # else assume annual
            obj, created = AnnualData.objects.update_or_create(
                indicator=target_indicator,
                for_datapoint=datapoint,
                defaults={'performance': performance, 'is_verified': True}
            )
            if created: result['created'] += 1
            else: result['updated'] += 1

        except Exception as e:
            result['skipped'] += 1
            result['errors'].append({'row': i, 'error': str(e)})

    return result


@api_view(['POST'])
@transaction.atomic
def submit_bulk_data_api(request):
    if not request.user.is_importer:
        return Response({'error': 'Access denied. Only importers can submit data.'}, status=status.HTTP_403_FORBIDDEN)

    data_file = request.FILES.get('data_file')
    if not data_file:
        return Response({'error': 'data_file is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Read rows from uploaded file (CSV or Excel)
        filename = (getattr(data_file, 'name', '') or '').lower()
        _, ext = os.path.splitext(filename)
        ext = ext.lower()

        def iter_rows_from_upload(fobj, extension):
            if extension == '.csv':
                # Read entire stream as text and parse via DictReader
                content = fobj.read()
                if isinstance(content, bytes):
                    text = content.decode('utf-8', errors='replace')
                else:
                    text = str(content)
                lines = text.splitlines()
                reader = csv.DictReader(lines)
                for r in reader:
                    yield r
            elif extension in ('.xls', '.xlsx'):
                if not openpyxl:
                    raise RuntimeError('Excel import requires openpyxl on the server. Please install openpyxl.')
                fobj.seek(0)
                wb = openpyxl.load_workbook(fobj, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.rows)
                if not rows:
                    return
                headers = [str(c.value).strip() if c.value is not None else '' for c in rows[0]]
                for row in rows[1:]:
                    obj = {}
                    for h, cell in zip(headers, row):
                        obj[h] = cell.value
                    yield obj
            else:
                raise RuntimeError('Unsupported file type. Accepts .csv, .xls, .xlsx')

        result = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}
        selected_indicator_ids = request.data.getlist('indicator_ids[]') or request.data.getlist('indicator_ids')
        allowed_indicator_codes = []
        if selected_indicator_ids:
            allowed_indicator_codes = list(Indicator.objects.filter(id__in=selected_indicator_ids).values_list('code', flat=True))

        processed_indicators = set()
        for i, raw in enumerate(iter_rows_from_upload(data_file, ext), start=1):
            try:
                row = { (str(k) or '').strip().lower(): v for k,v in raw.items() }
                
                # Check format
                is_annual_wide = any(k.isdigit() and len(k) == 4 for k in row.keys())
                is_period_wide = ('for_datapoint' in row) and ('for_quarter' in row or 'for_month' in row)

                if is_period_wide:
                    # Wide-Indicator format: for_datapoint, for_quarter/for_month, IND1, IND2...
                    control_cols = ['for_datapoint', 'for_quarter', 'for_month', 'indicator', 'title_eng', 'title_amh']
                    for col_name, val in row.items():
                        if col_name in control_cols or val in (None, ''):
                            continue
                        
                        # col_name is expected to be indicator code
                        if allowed_indicator_codes and col_name not in allowed_indicator_codes:
                            continue
                        
                        try:
                            ind_obj = Indicator.objects.get(code__iexact=col_name)
                            float(val)
                            processed_indicators.add(ind_obj)
                            result['created'] += 1
                        except Indicator.DoesNotExist:
                            pass
                        except Exception:
                            result['errors'].append({'row': i, 'error': f'Invalid value in column {col_name}: {val}'})
                    continue

                code = (row.get('indicator') or '').strip()
                if not code:
                    result['skipped'] += 1
                    result['errors'].append({'row': i, 'error': 'Missing indicator code'})
                    continue

                if allowed_indicator_codes and code not in allowed_indicator_codes:
                    result['skipped'] += 1
                    continue

                try:
                    indicator = Indicator.objects.get(code=code)
                except Indicator.DoesNotExist:
                    result['skipped'] += 1
                    result['errors'].append({'row': i, 'error': f'Unknown indicator code: {code}'})
                    continue

                if is_annual_wide:
                    # Annual Wide (Years as columns)
                    year_cols = [k for k in row.keys() if k.isdigit() and len(k) == 4]
                    for year in year_cols:
                        perf_raw = row.get(year)
                        if perf_raw in (None, ''):
                            continue
                        try:
                            float(perf_raw)
                            processed_indicators.add(indicator)
                            result['created'] += 1
                        except Exception:
                            result['errors'].append({'row': i, 'error': f'Invalid performance value for {year}: {perf_raw}'})
                else:
                    # Long format
                    year = row.get('year_ec') or row.get('year_gc')
                    perf_raw = row.get('performance') or row.get('value') or row.get('amount')
                    
                    if not year or perf_raw in (None, ''):
                        result['skipped'] += 1
                        result['errors'].append({'row': i, 'error': 'Missing required fields (year or performance or year-columns)'})
                        continue
                    
                    try:
                        float(perf_raw)
                        processed_indicators.add(indicator)
                        result['created'] += 1
                    except Exception:
                        result['skipped'] += 1
                        result['errors'].append({'row': i, 'error': f'Invalid performance value: {perf_raw}'})
                        continue

            except Exception as e:
                result['skipped'] += 1
                result['errors'].append({'row': i, 'error': str(e)})

        # Create a single DataSubmission record for the entire multiple-mode upload
        notes = request.data.get('notes', 'Bulk import (Multiple Mode)')
        DataSubmission.objects.create(
            indicator=None, # Multiple indicators in this file
            submitted_by=request.user,
            data_file=data_file,
            status='pending',
            notes=notes
        )

        return Response({
            'message': 'Bulk submission successful! A single pending submission record has been created for manager approval.',
            'total_indicators': len(processed_indicators),
            'rows_validated': result['created'],
            'rows_skipped': result['skipped'],
            'errors': result['errors']
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def decline_submission_api(request):
    """Decline a submission (indicator or data)"""
    submission_type = request.data.get('type')
    submission_id = request.data.get('id')
    
    if submission_type == 'indicator':
        submission = get_object_or_404(IndicatorSubmission, id=submission_id)
    elif submission_type == 'data':
        submission = get_object_or_404(DataSubmission, id=submission_id)
    else:
        return Response({'error': 'Invalid submission type'}, status=status.HTTP_400_BAD_REQUEST)
    
    submission.status = 'declined'
    submission.verified_by = request.user
    submission.verified_at = timezone.now()
    submission.save()
    # If this is an indicator submission, unmark the indicator as verified
    if submission_type == 'indicator' and hasattr(submission, 'indicator') and submission.indicator:
        try:
            ind = submission.indicator
            ind.is_verified = False
            ind.save(update_fields=['is_verified'])
        except Exception:
            pass
    
    if submission_type == 'indicator':
        serializer = IndicatorSubmissionSerializer(submission)
    else:
        serializer = DataSubmissionSerializer(submission)
    
    return Response(serializer.data)


@api_view(['GET'])
def indicators_list_api(request):
    """Get list of indicators for importer to submit data for"""
    category_id = request.GET.get('category_id')
    indicators = Indicator.objects.all().order_by('title_ENG')
    
    if category_id:
        indicators = indicators.filter(for_category__id=category_id)
        
    out = []
    for ind in indicators:
        cats = [c.name_ENG for c in ind.for_category.all()]
        out.append({
            'id': ind.id,
            'title_eng': ind.title_ENG,
            'title_amh': ind.title_AMH,
            'is_verified': getattr(ind, 'is_verified', False),
            'categories': cats,
            'code': ind.code,
        })
    return Response(out)


@api_view(['GET'])
def sample_template_api(request):
    """Return a downloadable sample Excel file for annual/quarter/month data submissions.

    The template matches the approval import expectations:
      - annual: columns [year_EC, performance]
      - quarter: columns [year_EC, quarter, performance]
      - monthly: columns [year_EC, month, performance]
    """
    kind = (request.GET.get('type') or request.GET.get('kind') or '').strip().lower()
    multiple = (request.GET.get('multiple') or '').strip().lower() in ('1', 'true', 'yes')
    category_id = request.GET.get('category_id')
    indicator_id = request.GET.get('indicator_id')
    
    if kind not in ('annual', 'quarter', 'monthly', 'weekly', 'daily'):
        return Response({'error': 'Invalid type. Use one of: annual, quarter, monthly, weekly, daily'}, status=status.HTTP_400_BAD_REQUEST)

    # Prepare pre-fill data
    prefill_indicators = []
    indicator_ids = request.GET.getlist('indicator_ids[]') or request.GET.getlist('indicator_ids') or ([request.GET.get('indicator_id')] if request.GET.get('indicator_id') else [])
    
    if multiple and indicator_ids:
        # User selected specific indicators for bulk template
        prefill_indicators = list(Indicator.objects.filter(id__in=indicator_ids).values('code', 'title_ENG', 'title_AMH'))
    elif multiple and category_id:
        # Fallback to all indicators in category if no specific ones selected
        prefill_indicators = list(Indicator.objects.filter(for_category__id=category_id).values('code', 'title_ENG', 'title_AMH'))
    elif not multiple and indicator_ids:
        # Single mode with one selection
        try:
            indicator = Indicator.objects.filter(id=indicator_ids[0]).values('code', 'title_ENG', 'title_AMH').first()
            if indicator:
                prefill_indicators = [indicator]
        except (Indicator.DoesNotExist, IndexError, ValueError):
            pass

    # Build sample dataset
    if kind == 'annual':
        headers = ['indicator', '2016', '2017', '2018', '2019']
        if prefill_indicators:
            rows = [(ind['code'] or ind['title_ENG'], '', '', '', '') for ind in prefill_indicators]
        else:
            rows = [('IND001', '123.45', '130.00', '', '')]
        filename = f"sample_{kind}_data.xlsx"
    elif kind == 'quarter':
        # Wide-indicator format: for_datapoint, for_quarter, IND1, IND2...
        headers = ['for_datapoint', 'for_quarter']
        if prefill_indicators:
            headers += [ind['code'] or ind['title_ENG'] for ind in prefill_indicators]
            rows = []
            for q in range(1, 5):
                row = ['2016', q] + ([''] * len(prefill_indicators))
                rows.append(row)
        else:
            headers += ['IND001', 'IND002']
            rows = [
                ('2017', 1, '43', '342'),
                ('2017', 2, '2', '243'),
                ('2017', 3, '12', '43'),
                ('2017', 4, '32', '43')
            ]
        filename = f"sample_{kind}_data.xlsx"
    elif kind == 'monthly':
        # Wide-indicator format: for_datapoint, for_month, IND1, IND2...
        headers = ['for_datapoint', 'for_month']
        if prefill_indicators:
            headers += [ind['code'] or ind['title_ENG'] for ind in prefill_indicators]
            rows = []
            for m in range(1, 13):
                row = ['2016', m] + ([''] * len(prefill_indicators))
                rows.append(row)
        else:
            headers += ['IND001', 'IND002']
            rows = [
                ('2017', m, '10', '20') for m in range(1, 13)
            ]
        filename = f"sample_{kind}_data.xlsx"
    elif kind == 'weekly':
        headers = ('indicator', 'title_eng', 'title_amh', 'year_EC', 'month', 'week', 'performance')
        if prefill_indicators:
            rows = []
            for ind in prefill_indicators:
                for m in range(1, 13):
                    for w in range(1, 6): # Up to 5 weeks
                        rows.append((ind['code'] or ind['title_ENG'], ind['title_ENG'], ind['title_AMH'], '2016', m, w, ''))
        else:
            rows = [('IND001', 'Sample English Title', 'የናሙና አማርኛ ርዕስ', '2016', 1, 1, 10.0)]
        filename = f"sample_{kind}_data.xlsx"
    elif kind == 'daily':
        headers = ('indicator', 'title_eng', 'title_amh', 'year_EC', 'month', 'day', 'performance')
        if prefill_indicators:
            rows = []
            for ind in prefill_indicators:
                # 1st day of each month for sample
                for m in range(1, 13):
                    rows.append((ind['code'] or ind['title_ENG'], ind['title_ENG'], ind['title_AMH'], '2016', m, 1, ''))
        else:
            rows = [('IND001', 'Sample English Title', 'የናሙና አማርኛ ርዕስ', '2016', 1, 1, 10.0)]
        filename = f"sample_{kind}_data.xlsx"
    else:
        # Fallback
        headers = ('indicator', 'title_eng', 'title_amh', 'year_EC', 'performance')
        rows = [('IND001', 'Sample Title', '', '2016', 0)]
        filename = f"sample_data.xlsx"

    dataset = tablib.Dataset(*rows, headers=headers)

    # Prefer xlsx output; fall back to csv if openpyxl not available
    try:
        binary = dataset.export('xlsx')
        resp = HttpResponse(binary, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    except Exception:
        csv_data = dataset.export('csv')
        resp = HttpResponse(csv_data, content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="{os.path.splitext(filename)[0]}.csv"'
        return resp


@api_view(['POST'])
def submit_indicator_api(request):
    """Submit a new indicator for approval"""
    if not request.user.is_importer:
        return Response({'error': 'Access denied. Only importers can submit indicators.'}, 
                       status=status.HTTP_403_FORBIDDEN)

    title_eng = request.data.get('title_eng')
    title_amh = request.data.get('title_amh')
    code = request.data.get('code')
    description = request.data.get('description')
    measurement_units = request.data.get('measurement_units')
    frequency = request.data.get('frequency')
    source = request.data.get('source')
    methodology = request.data.get('methodology')
    
    # New fields
    data_type = request.data.get('data_type')
    responsible_entity = request.data.get('responsible_entity')
    tags = request.data.get('tags')
    sdg_link = request.data.get('sdg_link')
    kpi_characteristics = request.data.get('kpi_characteristics')

    # Accept either a list or comma-separated string
    category_ids = request.data.get('category_ids') or request.data.get('category_id')

    if not title_eng or not category_ids:
        return Response({'error': 'Title (English) and category are required.'}, 
                       status=status.HTTP_400_BAD_REQUEST)

    # Normalize category_ids to a list of ints
    if isinstance(category_ids, str):
        # try as comma separated ids
        category_ids = [c.strip() for c in category_ids.split(',') if c.strip().isdigit()]
    if isinstance(category_ids, list):
        try:
            category_ids = [int(c) for c in category_ids]
        except ValueError:
            return Response({'error': 'Invalid category ids.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # validate categories
        categories = list(Category.objects.filter(id__in=category_ids))
        if len(categories) != len(category_ids):
            return Response({'error': 'One or more categories are invalid.'}, status=status.HTTP_400_BAD_REQUEST)

        # Validate code uniqueness if provided
        if code:
            if Indicator.objects.filter(code=code).exists():
                return Response({'error': f'Indicator with code "{code}" already exists.'}, 
                               status=status.HTTP_400_BAD_REQUEST)
        
        # create indicator without assigning m2m directly
        indicator = Indicator.objects.create(
            title_ENG=title_eng,
            title_AMH=title_amh,
            code=code if code else None,  # Use None instead of '' to avoid UNIQUE constraint with multiple empty strings
            description=description,
            measurement_units=measurement_units,
            frequency=frequency,
            source=source,
            methodology=methodology,
            data_type=data_type,
            responsible_entity=responsible_entity,
            tags=tags,
            sdg_link=sdg_link,
            kpi_characteristics=kpi_characteristics,
        )
        # assign categories (many-to-many)
        indicator.for_category.set(category_ids)

        # If code was not provided, generate it now that categories are assigned
        if not indicator.code:
            indicator.generate_code()
            indicator.save()

        # Create submission record
        IndicatorSubmission.objects.create(
            indicator=indicator,
            submitted_by=request.user,
            status='pending'
        )

        return Response({'message': 'Indicator submitted successfully', 'indicator_id': indicator.id}, 
                       status=status.HTTP_201_CREATED)
    except Category.DoesNotExist:
        return Response({'error': 'Invalid category'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def submit_data_api(request):
    """Submit data for an existing indicator"""
    if not request.user.is_importer:
        return Response({'error': 'Access denied. Only importers can submit data.'}, 
                       status=status.HTTP_403_FORBIDDEN)
    
    indicator_id = request.data.get('indicator_id')
    data_file = request.FILES.get('data_file')
    notes = request.data.get('notes', '')
    
    # If indicator_id is missing, but we have a data file, we can still proceed (it's effectively a bulk upload)
    if not indicator_id and not data_file:
        return Response({'error': 'Indicator ID or a data file is required.'}, 
                       status=status.HTTP_400_BAD_REQUEST)
    
    indicator = None
    if indicator_id:
        try:
            indicator = Indicator.objects.get(id=indicator_id)
        except Exception:
            return Response({'error': 'Invalid indicator ID.'}, status=status.HTTP_400_BAD_REQUEST)

    # Validate uploaded file format (read only header / first row) to avoid expensive loops
    if data_file:
        filename = (getattr(data_file, 'name', '') or '').lower()
        _, ext = os.path.splitext(filename)
        ext = ext.lower()

        # Helper to check required columns
        def _check_columns(cols):
            cols_l = [str(c).strip().lower() for c in cols if c is not None]
            has_year = any(x in cols_l for x in ('year_ec', 'year_gc', 'for_datapoint'))
            has_perf = any(x in cols_l for x in ('performance', 'value', 'amount'))
            has_period = any(x in cols_l for x in ('for_quarter', 'for_month'))
            # Also allow numeric headers (wide format)
            has_numeric_year = any(str(x).isdigit() and len(str(x)) == 4 for x in cols_l)
            # Valid if: (Standard Long) OR (Annual Wide) OR (Quarterly/Monthly Wide)
            return (has_year and has_perf) or has_numeric_year or (has_year and has_period)

        if ext == '.csv':
            try:
                # read a small chunk and extract first non-empty line as header
                data_file.seek(0)
                sample = data_file.read(8192)
                # ensure we have text
                if isinstance(sample, bytes):
                    sample_text = sample.decode('utf-8', errors='replace')
                else:
                    sample_text = str(sample)
                first_line = None
                for ln in sample_text.splitlines():
                    if ln.strip():
                        first_line = ln
                        break
                if not first_line:
                    return Response({'error': 'CSV file appears empty or malformed (no header found).'}, status=status.HTTP_400_BAD_REQUEST)
                # detect delimiter simply
                delimiter = ',' if first_line.count(',') >= first_line.count(';') else ';'
                import csv as _csv
                header = next(_csv.reader([first_line], delimiter=delimiter))
                if not _check_columns(header):
                    return Response({'error': 'CSV missing required columns. Required: year_EC/year_GC and performance/value/amount (case-insensitive).'}, status=status.HTTP_400_BAD_REQUEST)
            finally:
                try:
                    data_file.seek(0)
                except Exception:
                    pass
        elif ext in ('.xls', '.xlsx'):
            if not openpyxl:
                return Response({'error': 'Excel import requires openpyxl on the server. Please install openpyxl.'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                data_file.seek(0)
                wb = openpyxl.load_workbook(data_file, read_only=True, data_only=True)
                ws = wb.active
                # read header row only
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not header_row:
                    return Response({'error': 'Excel file appears empty or malformed (no header found).'}, status=status.HTTP_400_BAD_REQUEST)
                cols = [ (str(c).strip() if c is not None else '') for c in header_row ]
                if not _check_columns(cols):
                    return Response({'error': 'Excel missing required columns. Required: year_EC/year_GC and performance/value/amount (case-insensitive).'}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({'error': 'Failed to inspect Excel file header: ' + str(e)}, status=status.HTTP_400_BAD_REQUEST)
            finally:
                try:
                    data_file.seek(0)
                except Exception:
                    pass
        else:
            return Response({'error': 'Unsupported file type. Accepts .csv, .xls, .xlsx only.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Create submission record (file passed validations)
        submission = DataSubmission.objects.create(
            indicator=indicator,
            submitted_by=request.user,
            data_file=data_file,
            notes=notes,
            status='pending'
        )

        return Response({'message': 'Data submitted successfully', 'submission_id': submission.id}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def preview_existing_submission_api(request, submission_id: int):
    """
    Return a lightweight preview of an uploaded DataSubmission file.
    """
    try:
        submission = DataSubmission.objects.get(pk=submission_id)
    except DataSubmission.DoesNotExist:
        return Response({'error': 'Submission not found'}, status=status.HTTP_404_NOT_FOUND)

    ffield = submission.data_file
    if not ffield:
        return Response({'error': 'No data file attached to submission'}, status=status.HTTP_400_BAD_REQUEST)

    filename = (getattr(ffield, 'name', '') or '').lower()
    if filename.endswith('.csv'):
        rows = _preview_csv_file(ffield)
    elif filename.endswith(('.xls', '.xlsx')):
        if not openpyxl:
            return Response({'error': 'Excel preview requires openpyxl on the server.'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        rows = _preview_excel_file(ffield)
    else:
        return Response({'error': 'Unsupported file type for preview. Accepts .csv, .xls, .xlsx only.'},
                        status=status.HTTP_400_BAD_REQUEST)

    return Response({
        'id': submission.id,
        'name': getattr(ffield, 'name', ''),
        'size': getattr(ffield, 'size', 0),
        'rows': rows,
    }, status=status.HTTP_200_OK)


def _preview_csv_file(file_field, max_rows: int = 20):
    file_field.open('rb')
    wrapper = TextIOWrapper(file_field, encoding='utf-8', errors='replace')
    reader = csv.DictReader(wrapper)
    out = []
    for i, row in enumerate(reader):
        if i >= max_rows:
            break
        out.append({str(k): ('' if v is None else str(v)) for k, v in row.items()})
    file_field.close()
    return out


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def preview_data_submission_api(request):
    fobj = request.FILES.get('data_file')
    if not fobj:
        return Response({'error': 'data_file is required for preview.'},
                        status=status.HTTP_400_BAD_REQUEST)

    filename = (getattr(fobj, 'name', '') or '').lower()
    try:
        if filename.endswith('.csv'):
            rows_dict = _preview_csv_file(fobj, max_rows=50)
        elif filename.endswith(('.xls', '.xlsx')):
            if not openpyxl:
                return Response(
                    {'error': 'Excel preview requires openpyxl on the server.'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            rows_dict = _preview_excel_file(fobj, max_rows=50)
        else:
            return Response(
                {'error': 'Unsupported file type. Accepts .csv, .xls, .xlsx only.'},
                status=status.HTTP_400_BAD_REQUEST
            )
    except Exception as e:
        return Response({'error': f'Failed to parse file for preview: {e}'},
                        status=status.HTTP_400_BAD_REQUEST)

    if not rows_dict:
        preview = {'headers': [], 'rows': []}
        row_count = 0
    else:
        headers = list(rows_dict[0].keys())
        rows_matrix = []
        for row in rows_dict:
            rows_matrix.append([row.get(h, '') for h in headers])
        preview = {'headers': headers, 'rows': rows_matrix}
        row_count = len(rows_dict)

    payload = {
        'row_count': row_count,
        'preview': preview,
        'parse_errors': [],
        'invalid_rows': [],
        'warnings': [],
        'totals': {
            'new': row_count,
            'update': 0,
            'skip': 0,
        },
        'indicators_found': [],
        'row_statuses': [],
    }

    return Response(payload, status=status.HTTP_200_OK)


def _preview_excel_file(file_field, max_rows: int = 20):
    file_field.open('rb')
    wb = openpyxl.load_workbook(file_field, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.rows)
    if not rows:
        return []

    headers = [str(c.value).strip() if c.value is not None else '' for c in rows[0]]
    out = []
    for row in rows[1:max_rows + 1]:
        obj = {}
        for h, cell in zip(headers, row):
            obj[h] = '' if cell.value is None else str(cell.value)
        out.append(obj)
    file_field.close()
    return out


class AnnualSidebarList(APIView):
    PAGE_SIZE = 10

    def get(self, request):
        # Accept 1-based page from frontend; convert to 0-based slice index
        page = max(int(request.GET.get("page", 1)), 1)
        datapoints = DataPoint.objects.order_by('-year_EC')
        total = datapoints.count()
        start = (page - 1) * self.PAGE_SIZE
        slice_items = list(datapoints[start:start + self.PAGE_SIZE])

        payload = [
            {"id": dp.id, "year_ec": dp.year_EC, "year_gc": dp.year_GC}
            for dp in slice_items
        ]

        total_pages = (total + self.PAGE_SIZE - 1) // self.PAGE_SIZE if self.PAGE_SIZE else 1
        next_url = None
        prev_url = None
        base = request.build_absolute_uri(request.path)
        if (start + self.PAGE_SIZE) < total:
            next_url = f"{base}?page={page+1}"
        if page > 1 and total > 0:
            prev_url = f"{base}?page={page-1}"

        return Response({
            "results": payload,
            "has_next": (start + self.PAGE_SIZE) < total,
            "has_prev": page > 1 and total > 0,
            "page": page,
            "total": total,
            "total_pages": total_pages,
            "next": next_url,
            "previous": prev_url,
        }, status=status.HTTP_200_OK)


class QuarterlySidebarList(APIView):
    PAGE_SIZE = 1  # one year per page

    def get(self, request):
        # Accept 1-based page parameter from frontend
        page = max(int(request.GET.get("page", 1)), 1)

        datapoints = DataPoint.objects.order_by('-year_EC')
        total_years = datapoints.count()

        # select datapoint by 1-based page index
        idx = page - 1
        try:
            datapoint = datapoints[idx]
        except IndexError:
            datapoint = None

        payload = []
        if datapoint:
            for quarter in self._get_quarters():
                payload.append({
                    "id": f"{datapoint.year_EC}-Q{quarter['number']}",
                    "title_ENG": quarter['title_ENG'],
                    "title_AMH": quarter['title_AMH'],
                    "year_ec": datapoint.year_EC,
                    "year_gc": datapoint.year_GC,
                    "quarter_number": quarter['number'],
                })

        return Response({
            "results": payload,
            "has_next": page < total_years,
            "has_prev": page > 1,
            "page": page,
            "total": total_years,
            "total_pages": total_years,
            "next": (request.build_absolute_uri(request.path) + f"?page={page+1}") if page < total_years else None,
            "previous": (request.build_absolute_uri(request.path) + f"?page={page-1}") if page > 1 else None,
        })

    def _get_quarters(self):
        qs = Quarter.objects.order_by('number')
        if qs.exists():
            return [
                {"number": q.number,
                 "title_ENG": q.title_ENG,
                 "title_AMH": q.title_AMH} for q in qs
            ]
        return [{"number": i, "title_ENG": f"Q{i}", "title_AMH": f"Q{i}"} for i in range(1, 4)]


class MonthlySidebarList(APIView):
    PAGE_SIZE = 1  # one year per page

    def get(self, request):
        # Accept 1-based page parameter from frontend
        page = max(int(request.GET.get("page", 1)), 1)

        datapoints = DataPoint.objects.order_by('-year_EC')
        total_years = datapoints.count()

        idx = page - 1
        try:
            datapoint = datapoints[idx]
        except IndexError:
            datapoint = None

        payload = []
        if datapoint:
            for month in self._get_months():
                payload.append({
                    "id": f"{datapoint.year_EC}-M{month['number']}",
                    "month_ENG": month['month_ENG'],
                    "month_AMH": month['month_AMH'],
                    "year_ec": datapoint.year_EC,
                    "year_gc": datapoint.year_GC,
                    "month_number": month['number'],
                })

        return Response({
            "results": payload,
            "has_next": page < total_years,
            "has_prev": page > 1,
            "page": page,
            "total": total_years,
            "total_pages": total_years,
            "next": (request.build_absolute_uri(request.path) + f"?page={page+1}") if page < total_years else None,
            "previous": (request.build_absolute_uri(request.path) + f"?page={page-1}") if page > 1 else None,
        })

    def _get_months(self):
        qs = Month.objects.order_by('number')
        if qs.exists():
            return [{
                "number": m.number,
                "month_ENG": m.month_ENG,
                "month_AMH": m.month_AMH,
            } for m in qs]

        # fallback
        return [{"number": i, "month_ENG": f"Month {i}", "month_AMH": f"Month {i}"} for i in range(1, 13)]


class WeeklySidebarList(APIView):
    PAGE_SIZE = 5  # show 5 weeks per page

    def get(self, request):
        page = max(int(request.GET.get("page", 1)), 1)
        # Filter dates only for selected indicators if provided
        indicator_ids = request.GET.get('ids', '').split(',')
        indicator_ids = [i for i in indicator_ids if i.strip()]

        unique_dates_qs = KPIRecord.objects.filter(record_type='weekly', indicator__for_category__topic__is_initiative=False)
        if indicator_ids:
            unique_dates_qs = unique_dates_qs.filter(indicator_id__in=indicator_ids)
        
        unique_dates = unique_dates_qs.order_by('date').values_list('date', flat=True).distinct()

        seen_weeks = set()
        unique_week_items = []

        # Helper to calculate week grouping key
        def get_ethio_week_key(greg_date):
            gregorian_date = EthDate(greg_date.day, greg_date.month, greg_date.year)
            eth_date = to_ethiopian(gregorian_date)
            week = ((eth_date.day - 1) // 7) + 1
            week = min(week, 5)
            return (eth_date.year, eth_date.month, week)

        for d in unique_dates:
            key = get_ethio_week_key(d)
            if key not in seen_weeks:
                seen_weeks.add(key)
                unique_week_items.append({
                    'date': d,
                    'year_ec': key[0],
                    'month_ec': key[1],
                    'week_ec': key[2]
                })

        total = len(unique_week_items)
        start = (page - 1) * self.PAGE_SIZE
        page_items = unique_week_items[start:start + self.PAGE_SIZE]

        # Format for payload
        payload = []
        for item in page_items:
            rec_qs = KPIRecord.objects.filter(record_type='weekly', date=item['date'], indicator__for_category__topic__is_initiative=False)
            if indicator_ids:
                rec_qs = rec_qs.filter(indicator_id__in=indicator_ids)
            rec = rec_qs.first()
            if not rec:
                continue

            month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            # item['date'] is a date object
            month_gc_idx = item['date'].month - 1
            month_name = month_names[month_gc_idx]
            
            # Format: "Week X (Month) - Year" (Using GC Year to match Month Name)
            label = f"Week {item['week_ec']} ({month_name}) - {item['date'].year}"

            payload.append({
                "id": rec.id,
                "date": item['date'].isoformat(),
                "ethio_date": rec.ethio_date, # Should match our calc
                "week": item['week_ec'],
                "month_number": item['month_ec'],
                "month_name": month_name,
                "label": label,
                "year_ec": item['year_ec'],
                "year_gc": f"{item['date'].year}",
            })

        return Response({
            "results": payload,
            "has_next": (start + self.PAGE_SIZE) < total,
            "has_prev": page > 1,
            "page": page,
            "total": total,
            "total_pages": (total + self.PAGE_SIZE - 1) // self.PAGE_SIZE,
        })


class DailySidebarList(APIView):
    PAGE_SIZE = 10  # show 10 days per page

    def get(self, request):
        page = max(int(request.GET.get("page", 1)), 1)
        # Filter dates only for selected indicators if provided
        indicator_ids = request.GET.get('ids', '').split(',')
        indicator_ids = [i for i in indicator_ids if i.strip()]

        # Get all daily KPI records ordered by date (most recent first)
        qs = KPIRecord.objects.filter(record_type='daily', indicator__for_category__topic__is_initiative=False)
        if indicator_ids:
            qs = qs.filter(indicator_id__in=indicator_ids)
        
        unique_dates = list(qs.order_by('-date').values_list('date', flat=True).distinct())
        total = len(unique_dates)

        start = (page - 1) * self.PAGE_SIZE
        page_dates = unique_dates[start:start + self.PAGE_SIZE]
        slice_items = []
        for d in page_dates:
            rec = qs.filter(date=d).first()
            if rec:
                slice_items.append(rec)

        # Format each daily record for sidebar
        payload = []
        for r in slice_items:
            # Format Gregorian date
            greg_date_str = r.date.strftime("%b %d, %Y")  # e.g., "Dec 04, 2024"
            
            # Ethiopian date is already formatted in the model property
            ethio_date_str = r.ethio_date if r.ethio_date else ""
            
            payload.append({
                "id": r.id,
                "date": r.date.isoformat(),
                "greg_date_formatted": greg_date_str,
                "ethio_date": ethio_date_str,
                "label": f"{greg_date_str} ({ethio_date_str})",
                "performance": float(r.performance) if r.performance else None,
                "target": float(r.target) if r.target else None,
            })

        return Response({
            "results": payload,
            "has_next": (start + self.PAGE_SIZE) < total,
            "has_prev": page > 1,
            "page": page,
            "total": total,
            "total_pages": (total + self.PAGE_SIZE - 1) // self.PAGE_SIZE,
        })

# Category Assignments

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_category_assignment_api(request):
    manager_id = request.data.get('manager_id')
    category_id = request.data.get('category_id')

    if not manager_id:
        return Response({'error': 'Manager ID is required.'}, status=status.HTTP_400_BAD_REQUEST)

    if not category_id:
        return Response({'error': 'Category ID is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        manager_id = int(manager_id)
        category_id = int(category_id)
    except (ValueError, TypeError):
        return Response({'error': 'Invalid manager or category ID.'}, status=status.HTTP_400_BAD_REQUEST)

    # Validate manager
    try:
        manager = CustomUser.objects.get(id=manager_id)
        if not manager.is_category_manager:
            return Response(
                {'error': 'Selected user is not a category manager.'},
                status=status.HTTP_400_BAD_REQUEST
            )
    except CustomUser.DoesNotExist:
        return Response({'error': 'Manager not found.'}, status=status.HTTP_404_NOT_FOUND)

    # Validate category
    try:
        category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return Response({'error': 'Category not found.'}, status=status.HTTP_404_NOT_FOUND)

    # ✅ Category can have ONLY ONE manager
    if CategoryAssignment.objects.filter(category_id=category_id).exists():
        return Response(
            {'error': 'This category is already assigned to a manager.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    assignment = CategoryAssignment.objects.create(
        manager_id=manager_id,
        category_id=category_id
    )

    data = {
        'id': assignment.id,
        'manager_details': {
            'id': assignment.manager.id,
            'full_name': assignment.manager.get_full_name() or assignment.manager.username,
            'email': assignment.manager.email,
            'is_active': assignment.manager.is_active,
            'last_login': assignment.manager.last_login.isoformat() if assignment.manager.last_login else None,
        },
        'category_details': {
            'id': assignment.category.id,
            'name_eng': assignment.category.name_ENG,
            'indicator_count': assignment.category.indicators.count(),
            'subcategory_count': assignment.category.subcategories.count()
            if hasattr(assignment.category, 'subcategories') else 0,
            'topic_title': assignment.category.topic.title_ENG if assignment.category.topic else '',
        },
    }

    return Response(data, status=status.HTTP_201_CREATED)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_category_assignment_api(request, pk):
    try:
        assignment = CategoryAssignment.objects.get(pk=pk)
    except CategoryAssignment.DoesNotExist:
        return Response({'error': 'Assignment not found'}, status=status.HTTP_404_NOT_FOUND)

    manager_id = request.data.get('manager_id')

    if not manager_id:
        return Response({'error': 'Manager ID is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        manager_id = int(manager_id)
    except (ValueError, TypeError):
        return Response({'error': 'Invalid manager ID.'}, status=status.HTTP_400_BAD_REQUEST)

    # Validate manager
    try:
        manager = CustomUser.objects.get(id=manager_id)
        if not manager.is_category_manager:
            return Response(
                {'error': 'Selected user is not a category manager.'},
                status=status.HTTP_400_BAD_REQUEST
            )
    except CustomUser.DoesNotExist:
        return Response({'error': 'Manager not found.'}, status=status.HTTP_404_NOT_FOUND)

    # ✅ NO restriction on manager owning multiple categories
    assignment.manager_id = manager_id
    assignment.save(update_fields=['manager'])

    data = {
        'id': assignment.id,
        'manager_details': {
            'id': assignment.manager.id,
            'full_name': assignment.manager.get_full_name() or assignment.manager.username,
            'email': assignment.manager.email,
            'is_active': assignment.manager.is_active,
            'last_login': assignment.manager.last_login.isoformat() if assignment.manager.last_login else None,
        },
        'category_details': {
            'id': assignment.category.id,
            'name_eng': assignment.category.name_ENG,
            'indicator_count': assignment.category.indicators.count(),
            'subcategory_count': assignment.category.subcategories.count()
            if hasattr(assignment.category, 'subcategories') else 0,
            'topic_title': assignment.category.topic.title_ENG if assignment.category.topic else '',
        },
    }

    return Response(data, status=status.HTTP_200_OK)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_category_assignment_api(request, pk):
    try:
        assignment = CategoryAssignment.objects.get(pk=pk)
        category = assignment.category
        assignment.delete()

        data = {
            'success': True,
            'message': 'Assignment deleted successfully',
            'category': {
                'id': category.id,
                'name_ENG': category.name_ENG,
                'indicator_count': category.indicators.count()
            }
        }
        return Response(data, status=status.HTTP_200_OK)
    except CategoryAssignment.DoesNotExist:
        return Response({'error': 'Assignment not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': f'Failed to delete assignment: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def review_pending_data(request):
    """
    Fetch all unverified data points from AnnualData, QuarterData, MonthData, and KPIRecord.
    """
    if not (request.user.is_category_manager or request.user.is_superuser):
        return Response({'error': 'Unauthorized'}, status=403)

    results = []
    assigned_categories = []
    if not request.user.is_superuser:
        assigned_categories = list(CategoryAssignment.objects.filter(manager=request.user, category__topic__is_initiative=False).values_list('category_id', flat=True))

    # 1. Annual Data
    annuals = AnnualData.objects.filter(is_verified=False, indicator__for_category__topic__is_initiative=False).select_related('indicator', 'for_datapoint')
    if not request.user.is_superuser:
        annuals = annuals.filter(indicator__for_category__in=assigned_categories).distinct()
    
    for a in annuals:
        results.append({
            'id': a.id,
            'type': 'annual',
            'indicator_title': a.indicator.title_ENG if a.indicator else 'Unknown',
            'period_display': f"{a.for_datapoint.year_EC} EC" if a.for_datapoint else "Unknown Year",
            'value': a.performance
        })

    # 2. Quarterly Data
    quarters = QuarterData.objects.filter(is_verified=False, indicator__for_category__topic__is_initiative=False).select_related('indicator', 'for_quarter', 'for_datapoint')
    if not request.user.is_superuser:
        quarters = quarters.filter(indicator__for_category__in=assigned_categories).distinct()

    for q in quarters:
        q_title = q.for_quarter.title_ENG if q.for_quarter else f"Q{q.for_quarter_number or '?'}"
        year = q.for_datapoint.year_EC if q.for_datapoint else "?"
        results.append({
            'id': q.id,
            'type': 'quarterly',
            'indicator_title': q.indicator.title_ENG if q.indicator else 'Unknown',
            'period_display': f"{year} EC - {q_title}",
            'value': q.performance
        })

    # 3. Monthly Data
    months = MonthData.objects.filter(is_verified=False, indicator__for_category__topic__is_initiative=False).select_related('indicator', 'for_month', 'for_datapoint')
    if not request.user.is_superuser:
        months = months.filter(indicator__for_category__in=assigned_categories).distinct()

    for m in months:
        m_name = m.for_month.month_ENG if m.for_month else "?"
        year = m.for_datapoint.year_EC if m.for_datapoint else "?"
        results.append({
            'id': m.id,
            'type': 'monthly',
            'indicator_title': m.indicator.title_ENG if m.indicator else 'Unknown',
            'period_display': f"{year} EC - {m_name}",
            'value': m.performance
        })

    # 4. KPI Records (Weekly/Daily)
    kpis = KPIRecord.objects.filter(is_verified=False, indicator__for_category__topic__is_initiative=False).select_related('indicator')
    if not request.user.is_superuser:
        kpis = kpis.filter(indicator__for_category__in=assigned_categories).distinct()

    for k in kpis:
        p_disp = str(k.date)
        if k.record_type == 'weekly':
            p_disp = f"Weekly: {k.date}"
        elif k.record_type == 'daily':
            p_disp = f"Daily: {k.date}"
        
        results.append({
            'id': k.id,
            'type': k.record_type, 
            'indicator_title': k.indicator.title_ENG if k.indicator else 'Unknown',
            'period_display': p_disp,
            'value': k.performance
        })

    return Response({'results': results})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_pending_data(request):
    if not (request.user.is_category_manager or request.user.is_superuser):
        return Response({'error': 'Unauthorized'}, status=403)
    
    try:
        data_type = request.data.get('type')
        data_id = request.data.get('id')
        
        if not data_type or not data_id:
            return Response({'error': 'Missing type or id'}, status=400)
        
        obj = None
        if data_type == 'annual':
            obj = AnnualData.objects.get(id=data_id)
        elif data_type == 'quarterly':
            obj = QuarterData.objects.get(id=data_id)
        elif data_type == 'monthly':
            obj = MonthData.objects.get(id=data_id)
        elif data_type == 'weekly' or data_type == 'daily':
            obj = KPIRecord.objects.get(id=data_id)
        else:
            return Response({'error': 'Invalid type'}, status=400)
        
        obj.is_verified = True
        obj.is_seen = True # Clear the blue dot after approval
        obj.save()
        return Response({'status': 'approved'})

    except (AnnualData.DoesNotExist, QuarterData.DoesNotExist, MonthData.DoesNotExist, KPIRecord.DoesNotExist):
        return Response({'error': 'Data record not found'}, status=404)
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def decline_pending_data(request):
    """Decline (delete) specific table data edits"""
    if not (request.user.is_category_manager or request.user.is_superuser):
        return Response({'error': 'Unauthorized'}, status=403)
    
    try:
        data_type = request.data.get('type')
        data_id = request.data.get('id')
        
        if not data_type or not data_id:
            return Response({'error': 'Missing type or id'}, status=400)
        
        if data_type == 'annual':
            AnnualData.objects.filter(id=data_id).delete()
        elif data_type == 'quarterly':
            QuarterData.objects.filter(id=data_id).delete()
        elif data_type == 'monthly':
            MonthData.objects.filter(id=data_id).delete()
        elif data_type in ('weekly', 'daily'):
            KPIRecord.objects.filter(id=data_id).delete()
        else:
            return Response({'error': 'Invalid type'}, status=400)
        
        return Response({'status': 'declined'})
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approve_all_table_data_api(request):
    """Approve all pending table data submissions for assigned categories"""
    if not (request.user.is_category_manager or request.user.is_superuser):
        return Response({'error': 'Unauthorized'}, status=403)
    
    try:
        q_filter = Q(is_verified=False)
        if not request.user.is_superuser:
            assigned_categories = CategoryAssignment.objects.filter(manager=request.user, category__topic__is_initiative=False).values_list('category_id', flat=True)
            q_filter &= Q(indicator__for_category__in=assigned_categories)
        
        q_filter &= Q(indicator__for_category__topic__is_initiative=False)

        AnnualData.objects.filter(q_filter).update(is_verified=True, is_seen=True)
        QuarterData.objects.filter(q_filter).update(is_verified=True, is_seen=True)
        MonthData.objects.filter(q_filter).update(is_verified=True, is_seen=True)
        # KPIRecord filter doesn't support indicator__for_category directly in the same way if indicator is null, 
        # but we follow the same pattern
        KPIRecord.objects.filter(q_filter).update(is_verified=True, is_seen=True)

        return Response({'status': 'success', 'message': 'All pending table data approved'})
    except Exception as e:
        return Response({'error': str(e)}, status=500)